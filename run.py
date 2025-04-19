import openpyxl
import re
from datetime import datetime
import os
import logging # Using logging for better feedback

# --- Configuration ---
INPUT_DIR = "input"
OUTPUT_TEMPLATE = "output/template.xlsx"
OUTPUT_FILLED = "output/output_filled.xlsx"
START_ROW_TEMPLATE = 7  # Starting row in the output template
DATA_BLOCK_START_ROW = 19 # First row of the data block in input files
DATA_BLOCK_END_ROW = 1000   # Last row of the data block in input files

# --- Setup Logging ---
# Use INFO for general progress, WARNING for recoverable issues, ERROR for failures
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

# --- Helper Function ---
def prep_etd(data):
    """
    Parses the ETD string (e.g., "2024.08.08.(AIR)") to extract date and BA.
    Returns formatted date (DD-Mon) and BA string, or None, None on failure.
    """
    if not isinstance(data, str):
        logging.warning(f"ETD 데이터가 문자열이 아님: {data}. 처리할 수 없음.")
        return None, None
    # Regex to capture date (YYYY.MM.DD) and BA (inside parentheses)
    # Allows potential trailing dot after date
    match = re.search(r"([0-9]{4}\.[0-9]{1,2}\.[0-9]{1,2})\.?\s*\(([^)]+)\)", data)
    if not match:
        # Try alternative without trailing dot requirement for flexibility
        match = re.search(r"([0-9]{4}\.[0-9]{1,2}\.[0-9]{1,2})\s*\(([^)]+)\)", data)
        if not match:
            logging.warning(f"ETD 데이터 형식 불일치: '{data}'. 날짜/BA 추출 불가.")
            return None, None
    try:
        date_str_raw = match.group(1)
        ba_str = match.group(2)
        # Attempt to parse date, handle potential errors
        date_obj = datetime.strptime(date_str_raw, "%Y.%m.%d")
        date_str_formatted = date_obj.strftime("%d-%b")
        return date_str_formatted, ba_str
    except ValueError:
        logging.warning(f"ETD 날짜 형식 오류: '{date_str_raw}' in '{data}'. 날짜 변환 불가.")
        # Decide if you want to return BA even if date fails, or None, None
        return None, ba_str # Returning BA for now, date will be None
    except Exception as e:
        logging.error(f"prep_etd 처리 중 예외 발생 ('{data}'): {e}")
        return None, None

# --- Main Processing Logic ---

# Create output directory if it doesn't exist
output_dir = os.path.dirname(OUTPUT_FILLED)
if output_dir and not os.path.exists(output_dir):
    try:
        os.makedirs(output_dir)
        logging.info(f"'{output_dir}' 디렉토리를 생성했습니다.")
    except OSError as e:
        logging.error(f"출력 디렉토리 '{output_dir}' 생성 실패: {e}")
        exit() # Cannot proceed without output directory

# Load template workbook
try:
    template_wb = openpyxl.load_workbook(OUTPUT_TEMPLATE)
    template_sheet = template_wb.active
    logging.info(f"템플릿 파일 '{OUTPUT_TEMPLATE}' 로드 완료.")
except FileNotFoundError:
    logging.error(f"치명적 오류: 템플릿 파일 '{OUTPUT_TEMPLATE}'을 찾을 수 없습니다. 프로그램을 종료합니다.")
    exit()
except Exception as e:
    logging.error(f"치명적 오류: 템플릿 파일 '{OUTPUT_TEMPLATE}' 로드 중 오류 발생: {e}")
    exit()


# Find input files (excluding temporary Excel files starting with ~)
if not os.path.exists(INPUT_DIR):
    logging.error(f"치명적 오류: 입력 디렉토리 '{INPUT_DIR}'을 찾을 수 없습니다. 프로그램을 종료합니다.")
    exit()
try:
    # Ensure correct listing and filtering
    all_files_in_dir = os.listdir(INPUT_DIR)
    input_files = [
        f for f in all_files_in_dir
        if f.endswith('.xlsx')
        and not f.startswith('~')
        and os.path.isfile(os.path.join(INPUT_DIR, f)) # Make sure it's a file
    ]
    if not input_files:
        logging.warning(f"입력 디렉토리 '{INPUT_DIR}'에서 처리할 .xlsx 파일을 찾을 수 없습니다.")
        # No need to exit, just won't process anything
    else:
        logging.info(f"'{INPUT_DIR}' 디렉토리에서 {len(input_files)}개의 .xlsx 파일 발견.")

except Exception as e:
    logging.error(f"입력 디렉토리 '{INPUT_DIR}' 검색 중 오류 발생: {e}")
    exit()


# --- Optional Test override ---
# Uncomment the following lines to use a specific list for testing
# test_files_list = [ 'ETD08.08 AIR JCON 선적서류 JB24FW-M34.xlsx', 'ETD08.08 TRUCK JCON 선적서류 JB24FW-M32.xlsx', 'ETD08.10 BOAT JCON 선적서류 JB24FW-M34.xlsx', 'ETD08.12 EXPRESS JCON 선적서류 JB24FW-M104-2.xlsx' ,'ETD08.15 AIR JCON 선적서류 JB24FW-M32.xlsx', 'ETD08.10 BOAT JCON 선적서류 JB24FW-M34.xlsx', 'ETD08.19 TRUCK JCON 선적서류 JB24FW-M36.xlsx', 'ETD08.21 TRUCK JCON 선적서류 JB24FW-M36.xlsx', 'ETD08.22 AIR JCON 선적서류 JB24FW-M35.xlsx', 'ETD08.22 BOAT JCON 선적서류 JB24FW-M36.xlsx']
# input_files = [f for f in test_files_list if os.path.exists(os.path.join(INPUT_DIR, f))]
# missing_test_files = set(test_files_list) - set(input_files)
# if missing_test_files:
#      logging.warning(f"테스트 파일 목록 중 다음 파일들을 찾을 수 없어 제외합니다: {', '.join(missing_test_files)}")
# logging.info(f"테스트용 파일 목록 사용 (총 {len(input_files)}개): {input_files}")
# --- End Test override ---


current_output_row = START_ROW_TEMPLATE # Initialize the row counter for the output sheet
processed_files_count = 0

for filename in sorted(input_files):
    logging.info(f"--- 파일 처리 시작: {filename} ---")
    filepath = os.path.join(INPUT_DIR, filename)
    file_had_valid_data = False # Flag per file

    try:
        # Load the input workbook, getting actual values (not formulas)
        # Use read_only for potential speedup if not modifying input
        workbook = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        sheetnames = workbook.sheetnames

        if not sheetnames:
            logging.warning(f"파일 '{filename}'에 시트가 없습니다. 건너뜁니다.")
            workbook.close()
            continue

        # Assume data is always on the *first* sheet
        sheet = workbook[sheetnames[0]]
        logging.info(f"'{filename}' 파일의 첫 번째 시트 ('{sheet.title}') 사용.")

        # --- Extract common data (once per file) ---
        # Assume the *last* sheet name is the ORDER_NO
        ORDER_NO = sheetnames[-1] if sheetnames else None # Handle case with no sheets gracefully
        if ORDER_NO is None:
            logging.warning(f"파일 '{filename}'에서 ORDER_NO (마지막 시트 이름)를 추출할 수 없습니다. 건너뜁니다.")
            workbook.close()
            continue

        etd_data = sheet["I11"].value
        # Assume REMARK is always in A22 for the entire file (as per original code before keyword search)
        # If you want the keyword search back, uncomment the relevant block from the previous version
        REMARK = sheet["A22"].value

        ETD, BA = None, None # Initialize ETD and BA
        if etd_data is None:
            logging.warning(f"  I11 셀이 비어 있음. ETD/BA 정보 없음.")
        else:
            ETD, BA = prep_etd(str(etd_data)) # Ensure input is string
            # Log only if extraction failed or gave partial result
            if ETD is None and BA is None:
                 logging.warning(f"  ETD/BA ('{etd_data}') 처리 실패.")
            elif ETD is None:
                logging.warning(f"  ETD/BA ('{etd_data}') 처리: ETD 추출 실패, BA: {BA}")
            elif BA is None:
                 logging.warning(f"  ETD/BA ('{etd_data}') 처리: BA 추출 실패, ETD: {ETD}")
            # No need to log success every time unless debugging
            # else:
            #     logging.info(f"  추출된 ETD: {ETD}, BA: {BA}")


        logging.info(f"  추출된 공통 데이터 - ORDER_NO: {ORDER_NO}, REMARK: {REMARK}")

        # --- Loop through the potential data rows (19-28) in the input sheet ---
        data_found_in_block = False
        for r in range(DATA_BLOCK_START_ROW, DATA_BLOCK_END_ROW + 1):
            # Read data from the current row in the block
            # Use .cell(row, column).value for potentially better robustness
            DETAIL = sheet.cell(row=r, column=7).value # Column G
            COLOR = sheet.cell(row=r, column=8).value # Column H
            PRICE = sheet.cell(row=r, column=9).value # Column I
            SQNTY = sheet.cell(row=r, column=10).value # Column J

            # --- Condition 1: Check if row has *any* meaningful data ---
            # Process only if at least one key cell has a value that's not just whitespace
            has_data = any(val is not None and (not isinstance(val, str) or val.strip())
                           for val in [DETAIL, COLOR, PRICE, SQNTY])

            if has_data:
                data_found_in_block = True # Mark that we found *some* data in the block

                # --- Condition 2: Validate data types for PRICE and SQNTY ---
                # They must be numeric (int or float). None is not numeric.
                price_is_numeric = isinstance(PRICE, (int, float))
                sqnty_is_numeric = isinstance(SQNTY, (int, float))

                if price_is_numeric and sqnty_is_numeric:
                    # --- Data is valid, write to output ---
                    file_had_valid_data = True # Mark that this *file* contained valid data to write
                    logging.info(f"    - 입력 행 {r}: 유효 데이터 발견. 출력 행 {current_output_row}에 쓰는 중...")

                    template_sheet[f"B{current_output_row}"] = ETD
                    template_sheet[f"C{current_output_row}"] = BA
                    template_sheet[f"E{current_output_row}"] = ORDER_NO
                    template_sheet[f"F{current_output_row}"] = COLOR
                    template_sheet[f"G{current_output_row}"] = SQNTY
                    template_sheet[f"I{current_output_row}"] = PRICE
                    template_sheet[f"K{current_output_row}"] = DETAIL
                    template_sheet[f"P{current_output_row}"] = REMARK # Use the common remark

                    # --- Move to the next row in the output template ---
                    current_output_row += 1
                else:
                    # --- Data found, but invalid type, skip this row ---
                    logging.warning(f"    - 입력 행 {r}: 건너뜀. PRICE ('{PRICE}', type: {type(PRICE).__name__}) "
                                    f"또는 SQNTY ('{SQNTY}', type: {type(SQNTY).__name__})가 숫자가 아님.")
            # else:
                 # This row within the block was essentially empty or whitespace only
                 # logging.debug(f"    - 입력 행 {r}: 데이터 없음. 건너뜀.") # Optional: for debugging empty rows

        if not data_found_in_block:
             logging.warning(f"  파일 '{filename}'의 {DATA_BLOCK_START_ROW}-{DATA_BLOCK_END_ROW} 행 범위에서 데이터를 찾지 못했습니다.")
        elif not file_had_valid_data and data_found_in_block:
            # This case means we found rows with data, but none had valid numeric PRICE/SQNTY
             logging.warning(f"  파일 '{filename}'의 {DATA_BLOCK_START_ROW}-{DATA_BLOCK_END_ROW} 행 범위에서 데이터는 찾았으나, 유효한 숫자 타입의 PRICE/SQNTY를 가진 행이 없었습니다.")


        workbook.close() # Close the input workbook (important with read_only=True)
        processed_files_count += 1

    except FileNotFoundError:
        # This specific error is less likely here as we check existence before the loop, but keep for robustness
        logging.error(f"파일 '{filepath}'을(를) 처리 중 찾을 수 없습니다. 건너뜁니다.")
        continue
    except KeyError as e:
        # Error accessing sheet name or potentially cell (less likely with .cell)
        logging.error(f"파일 '{filename}' 처리 중 시트 이름 또는 셀 접근 오류 발생 ({e}). 건너뜁니다.")
        if 'workbook' in locals() and workbook: workbook.close() # Try to close if open
        continue
    except zipfile.BadZipFile:
         logging.error(f"파일 '{filename}'이(가) 유효한 Excel 파일이 아니거나 손상되었습니다. 건너뜁니다.")
         if 'workbook' in locals() and workbook: workbook.close()
         continue
    except Exception as e:
        logging.error(f"파일 '{filename}' 처리 중 예기치 않은 오류 발생: {e}", exc_info=True) # Include traceback for debugging
        if 'workbook' in locals() and workbook: workbook.close() # Try to close if open
        continue # Continue with the next file

# --- Final Save ---
if current_output_row > START_ROW_TEMPLATE: # Only save if data was actually written
    logging.info(f"총 {processed_files_count}개 파일 처리 완료. {current_output_row - START_ROW_TEMPLATE}개의 데이터 행을 출력 파일에 기록 중...")
    try:
        template_wb.save(OUTPUT_FILLED)
        logging.info(f"✅ 모든 데이터 처리가 완료되어 '{OUTPUT_FILLED}'에 저장되었습니다.")
    except PermissionError:
         logging.error(f"치명적 오류: 파일 '{OUTPUT_FILLED}' 저장 권한이 없습니다. 파일이 다른 프로그램에서 열려있는지 확인하세요.")
    except Exception as e:
        logging.error(f"치명적 오류: 최종 파일 '{OUTPUT_FILLED}' 저장 중 오류 발생: {e}")
elif processed_files_count > 0:
    logging.warning(f"총 {processed_files_count}개 파일 처리 완료. 그러나 유효한 데이터 행이 없어 최종 출력 파일을 저장하지 않았습니다.")
else:
     logging.info("처리할 입력 파일이 없거나 모든 파일 처리 중 오류가 발생하여 최종 출력 파일을 저장하지 않았습니다.")